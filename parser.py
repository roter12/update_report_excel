#!/usr/bin/env python
_C='Sheet1'
_B='%d.%m.%Y'
_A=True
import re,os,sys,csv,glob,xlrd
from datetime import datetime,timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color,PatternFill,Font,Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
debug=False
log_file_path='*.rep'
rep_file_path='head report.xlsx'
LOG_DATE=datetime.now()
LOG_END_DATE=datetime.now()
LOG_START_DATE=datetime.now()
LOG_LAST_DATE=datetime.now()
arg_len=len(sys.argv)
if arg_len>3:debug=_A
if arg_len>2:rep_file_path=sys.argv[2]
if arg_len>1:log_file_path=sys.argv[1]
no_rep_sday=2
no_rep_eday=3
no_rep_no=4
no_rep_sak=7
no_rep_bei=8
no_rep_dat=10
no_log_rid=1
no_log_date=2
no_log_time=3
no_log_dev=4
no_log_cue=10
no_log_cuv=12
no_log_med=14
no_log_stat=19
log_files=[]
for file_path in glob.glob(log_file_path):log_files.append(file_path)
if len(log_files)==0:print("Can't read log file: "+log_file_path);exit(1)
if not os.path.exists(rep_file_path):print("Can't open report file: "+rep_file_path);exit(1)
def find_log(id,no_row):
	global LOG_DATE,LOG_START_DATE,LOG_LAST_DATE
	if type(id)==float:id=str(int(id))
	if debug==_A:print('Finding device',id,'...')
	D=datetime.now().date().replace(year=1999)
	for E in log_files:
		if debug==_A:print('   ','in',E,':',end='')
		H=csv.reader(open(E,'r'),delimiter='\t');F=0
		for A in H:
			if F==1:LOG_DATE=datetime.strptime(A[4].strip(),_B).date()
			if len(A)>=no_log_dev:
				B=A[no_log_dev]
				if type(B)==float:B=str(int(B))
				if B.lstrip('0')==id.lstrip('0'):
					if re.search('^[0-9\\/\\-\\.]+$',A[no_log_date]):
						C=datetime.strptime(A[no_log_date].strip(),_B).date()
						if LOG_DATE+timedelta(days=-2*4)<=C:
							if C>D:
								D=C
								if debug==_A:print('',A[no_log_rid-1],end='')
								LOG_START_DATE=LOG_DATE.replace(day=1);G=LOG_DATE.replace(day=28)+timedelta(days=4);LOG_LAST_DATE=G-timedelta(days=G.day);update_report(no_row,A)
			F+=1
		if debug==_A:print('')
	return[]
def update_report(no_row,row):
	K='%H:%M:%S';J='FFFFFF00';H='.';G='^[0-9\\,\\.]+$';F='%d/%m/%Y';D=',';B=no_row;A=row;global LOG_DATE,LOG_START_DATE,LOG_LAST_DATE;I=load_workbook(rep_file_path);C=I[_C];E=PatternFill(start_color=J,end_color=J,fill_type='solid');C.cell(B,no_rep_sday).value=LOG_START_DATE.strftime(F);C.cell(B,no_rep_eday).value=LOG_LAST_DATE.strftime(F);L=datetime.strptime(A[no_log_date].strip(),_B).date();M=datetime.strptime(A[no_log_time].strip(),K).time();C.cell(B,no_rep_dat).value=L.strftime(F)+' '+M.strftime(K)
	if re.search(G,A[no_log_stat]):
		if A[no_log_stat].count(D)==1:C.cell(B,no_rep_sak).value=float(A[no_log_stat].replace(D,H))
		else:C.cell(B,no_rep_sak).value=float(A[no_log_stat].replace(D,''))
	else:C.cell(B,no_rep_sak).value=A[no_log_stat];C.cell(B,no_rep_sak).fill=E
	if int(A[no_log_med])==2:
		if re.search(G,A[no_log_cue]):
			if A[no_log_cue].count(D)==1:C.cell(B,no_rep_bei).value=float(A[no_log_cue].replace(D,H))
			else:C.cell(B,no_rep_bei).value=float(A[no_log_cue].replace(D,''))
		else:C.cell(B,no_rep_bei).value=A[no_log_cue];C.cell(B,no_rep_bei).fill=E
	elif re.search(G,A[no_log_cuv]):
		if A[no_log_cuv].count(D)==1:C.cell(B,no_rep_bei).value=float(A[no_log_cuv].replace(D,H))
		else:C.cell(B,no_rep_bei).value=float(A[no_log_cuv].replace(D,''))
	else:C.cell(B,no_rep_bei).value=A[no_log_cuv];C.cell(B,no_rep_bei).fill=E
	if LOG_END_DATE.date().day<LOG_END_DATE.date().month:I.save(rep_file_path)
workbook=xlrd.open_workbook(rep_file_path)
worksheet=workbook.sheet_by_name(_C)
num_rows=worksheet.nrows-1
idx_row=0
while idx_row<num_rows:idx_row+=1;row=worksheet.row(idx_row);find_log(row[no_rep_no-1].value,idx_row+1)