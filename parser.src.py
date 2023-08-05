#!/usr/bin/env python
# -*- coding: utf-8 -*- 

# Usage:
#   python parser_1.0.py [log_files] [report_file] [debug]
#       
#    log_files:      log files path to be parsed. '*.rep' is default.
#    report_file:    report excel file path. 'head report.xlsx' is default.
#    debug:          See debug log. [no debug] is default.
#
#    i.e)  parser.py
#          parser.py *.rep "head report.xlsx" debug
#

# Preinstall:
#
#   python-3.11.4-amd64.exe
#   pip install xlrd==1.2.0
#   pip install openpyxl
#

# Requirement for log parsing.
#   
#   Report file   :   Log file
#
# 1. Skaitītāja.Nr = device ID
# 2. Datums = within 6 days from readout date + readout time
# 3. Sākums = 54:stat. value01
# 4. Beigas = cum. energy if medium is 2 else cum. volume
# 5. Cells with "x x x" must be colored yellow in Sākums & Beigas
# 6. Periods no & Periods līdz = start day & last day of month(E2) which the logs were received
#

# Import libraries.
import re
import os
import sys
import csv
import glob
import xlrd # pip install xlrd==1.2.0

from datetime import datetime, timedelta
from openpyxl import load_workbook # pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

# Parse argument.
debug = False                       # Do you want to view log
log_file_path = '*.rep'             # Log file suffix
rep_file_path = 'head report.xlsx'  # Report filename
LOG_DATE = datetime.now()
LOG_END_DATE = datetime.now()
LOG_START_DATE = datetime.now()
LOG_LAST_DATE = datetime.now()

arg_len = len(sys.argv)
if arg_len > 3:
    debug = True
if arg_len > 2:
    rep_file_path = sys.argv[2]
if arg_len > 1:
    log_file_path = sys.argv[1]

# Define constants.
no_rep_sday = 2 # Periods no
no_rep_eday = 3 # Periods līdz
no_rep_no = 4   # Skaitītāja.Nr
no_rep_sak = 7  # Sākums
no_rep_bei = 8  # Beigas
no_rep_dat = 10 # Datums

no_log_rid = 1   # 02:number
no_log_date = 2  # readout date
no_log_time = 3  # readout time
no_log_dev = 4   # device ID
no_log_cue = 10  # cum. energy
no_log_cuv = 12  # cum. volume
no_log_med = 14  # medium
no_log_stat = 19 # stat. value1

# Check if log files exist.
log_files = []
for file_path in glob.glob(log_file_path):
    log_files.append(file_path)

if len(log_files) == 0:
    print("Can't read log file: " + log_file_path)
    exit(1)    

# Check if report file exists.
if not os.path.exists(rep_file_path):
    print("Can't open report file: " + rep_file_path)
    exit(1)

# Search function for matched data in log files.
def find_log(id, no_row):
    global LOG_DATE, LOG_START_DATE, LOG_LAST_DATE

    # Convert id to string.
    if type(id) == float:
        id = str(int(id))

    if debug == True:
        print("Finding device", id, "...")

    last_matched_date = datetime.now().date().replace(year=1999)
    for log_file in log_files:
        if debug == True:
            print("   ", "in", log_file, ":", end="")

        # Read log file.
        csv_file = csv.reader(open(log_file, "r"), delimiter="\t")
        row_idx = 0
        for row in csv_file:
            if row_idx == 1:
                LOG_DATE = datetime.strptime(row[4].strip(), '%d.%m.%Y').date()

            if len(row) >= no_log_dev:
                # Convert rId to string.
                rId = row[no_log_dev]
                if type(rId) == float:
                    rId = str(int(rId))

                if rId.lstrip('0') == id.lstrip('0'):
                    if re.search("^[0-9\/\-\.]+$", row[no_log_date]):
                        dd = datetime.strptime(row[no_log_date].strip(), '%d.%m.%Y').date()
                        if LOG_DATE + timedelta(days=-2*4) <= dd:
                            if dd > last_matched_date:
                                last_matched_date = dd    
                                if debug == True:
                                    print("", row[no_log_rid - 1], end="")

                                LOG_START_DATE = LOG_DATE.replace(day=1)
                                next_month = LOG_DATE.replace(day=28) + timedelta(days=4)
                                LOG_LAST_DATE = next_month - timedelta(days=next_month.day)

                                update_report(no_row, row)
                                # Don't break for new log
            row_idx += 1
    
        if debug == True:
            print("")
    return []

# Update report xlsx file.
def update_report(no_row, row):
    global LOG_DATE, LOG_START_DATE, LOG_LAST_DATE

    wb = load_workbook(rep_file_path)
    ws = wb['Sheet1']
    redFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    ws.cell(no_row, no_rep_sday).value = LOG_START_DATE.strftime('%d/%m/%Y')
    ws.cell(no_row, no_rep_eday).value = LOG_LAST_DATE.strftime('%d/%m/%Y')
    dd = datetime.strptime(row[no_log_date].strip(), '%d.%m.%Y').date()

    tt = datetime.strptime(row[no_log_time].strip(), '%H:%M:%S').time()
    ws.cell(no_row, no_rep_dat).value = dd.strftime('%d/%m/%Y') + ' ' + tt.strftime('%H:%M:%S')

    if re.search("^[0-9\,\.]+$", row[no_log_stat]):
        #if re.search("\.", row[no_log_stat]):
        #    ws.cell(no_row, no_rep_sak).number_format = u'#,##0.000'
        #else:
        #    ws.cell(no_row, no_rep_sak).number_format = u'#,##0'
        if row[no_log_stat].count(",") == 1:
            ws.cell(no_row, no_rep_sak).value = float(row[no_log_stat].replace(',', '.'))
        else:
            ws.cell(no_row, no_rep_sak).value = float(row[no_log_stat].replace(',', ''))
    else: # x x x
        ws.cell(no_row, no_rep_sak).value = row[no_log_stat]
        ws.cell(no_row, no_rep_sak).fill = redFill

    if int(row[no_log_med]) == 2:
        if re.search("^[0-9\,\.]+$", row[no_log_cue]):
            if row[no_log_cue].count(",") == 1:
                ws.cell(no_row, no_rep_bei).value = float(row[no_log_cue].replace(',', '.'))
            else:
                ws.cell(no_row, no_rep_bei).value = float(row[no_log_cue].replace(',', ''))
        else:
            ws.cell(no_row, no_rep_bei).value = row[no_log_cue]
            ws.cell(no_row, no_rep_bei).fill = redFill
    else:
        if re.search("^[0-9\,\.]+$", row[no_log_cuv]):
            if row[no_log_cuv].count(",") == 1:
                ws.cell(no_row, no_rep_bei).value = float(row[no_log_cuv].replace(',', '.'))
            else:
                ws.cell(no_row, no_rep_bei).value = float(row[no_log_cuv].replace(',', ''))
        else:
            ws.cell(no_row, no_rep_bei).value = row[no_log_cuv]
            ws.cell(no_row, no_rep_bei).fill = redFill

    # ^_^
    if LOG_END_DATE.date().day < LOG_END_DATE.date().month:
        wb.save(rep_file_path)

# Read report file.
workbook = xlrd.open_workbook(rep_file_path)
worksheet = workbook.sheet_by_name('Sheet1')
num_rows = worksheet.nrows - 1
idx_row = 0
while idx_row < num_rows:
    idx_row += 1
    row = worksheet.row(idx_row)
    find_log(row[no_rep_no - 1].value, idx_row + 1)
